import os
import json
import subprocess
import re
from docx import Document
from openpyxl import load_workbook

# Define el directorio base donde están las carpetas de los clientes
directorio_base = "C:\\Users\\E_noe\\Repos\\Creditos Avalfi"

# Diccionario para almacenar los datos por cliente
datos_por_cliente = []

# Función para extraer datos de archivos Word
def extraer_datos_word(ruta_archivo):
    doc = Document(ruta_archivo)
    datos = {}
    for para in doc.paragraphs:
        texto = para.text.strip().lower()  # Convertir a minúsculas para comparación insensible al caso
        
        # Comparar ambas versiones: con y sin tilde
        if "nombre del cliente:" in texto:
            datos["cliente"] = texto.split("nombre del cliente:")[1].strip()
            #
        elif "nº credito:" in texto or "nº crédito:" in texto:
            datos["n_credito"] = texto.split("nº credito:")[1].strip() if "nº credito:" in texto else texto.split("nº crédito:")[1].strip()
        elif "nombre del asesor:" in texto:
            nombre_asesor = texto.split("nombre del asesor:")[1].strip()
            #cambiar el nombre por un numero. Si contiene Karla = 1, Elvis = 2, Roberto = 3, Roxana = 4, Noe = 5
            if "karla" in nombre_asesor:
                datos["asesor"] = 1
            elif "elvis" in nombre_asesor:
                datos["asesor"] = 2
            elif "roberto" in nombre_asesor:
                datos["asesor"] = 3
            elif "roxana" in nombre_asesor:
                datos["asesor"] = 4
            elif "noe" in nombre_asesor:
                datos["asesor"] = 5
            else:
                datos["asesor"] = 0
        elif "fecha de comite:" in texto or "fecha de comité:" in texto:
            datos["fecha_comite"] = texto.split("fecha de comite:")[1].strip() if "fecha de comite:" in texto else texto.split("fecha de comité:")[1].strip()
            #convertir de 8/12/22 a 2022-12-08
            datos["fecha_comite"] = datos["fecha_comite"].replace("/", "-")
        elif "fecha de aprobacion:" in texto:
            datos["fecha_aprobacion"] = texto.split("fecha de aprobacion:")[1].strip()
            datos["fecha_aprobacion"] = datos["fecha_aprobacion"].replace("/", "-")
        elif "fecha de esembolso:" in texto:
            datos["fecha_desembolso"] = texto.split("fecha de desembolso:")[1].strip()
            datos["fecha_desembolso"] = datos["fecha_desembolso"].replace("/", "-")
        elif "monto aprobado:" in texto:
            datos["monto_aprobado"] = texto.split("monto aprobado:")[1].strip()
            #convertir a float y limpiar signo de dollar
            datos["monto_aprobado"] = float(re.sub(r'[^\d.]', '', datos["monto_aprobado"]))
        elif "forma de pago:" in texto:
            forma_pago = texto.split("forma de pago:")[1].strip()
            #forma de pago puede ser semanal, quincenal, mensual, etc. Usar 1 para semanal, 2 para quincenal, 3 para mensual
            if "semanal" in forma_pago:
                datos["forma_pago"] = 1
            elif "quincenal" in forma_pago:
                datos["forma_pago"] = 3
            elif "mensual" in forma_pago:
                datos["forma_pago"] = 4
            elif "trimestral" in forma_pago:
                datos["forma_pago"] = 5
            else:
                datos["forma_pago"] = 0
        elif "cuota:" in texto:
            datos["cuota"] = texto.split("cuota:")[1].strip()
        elif "no. de cuotas:" in texto:
            datos["no. de cuotas"] = texto.split("no. de cuotas:")[1].strip()
        elif "destino:" in texto:
            datos["destino"] = texto.split("destino:")[1].strip()
        elif "asesoria financiera:" in texto:
            datos["asesoria financiera"] = texto.split("asesoria financiera:")[1].strip()
        elif "iva:" in texto:
            datos["iva"] = texto.split("iva:")[1].strip()
            #limpiar signo de dollar y convertirlo en float
            datos["iva"] = float(re.sub(r'[^\d.]', '', datos["iva"]))
        elif "seguro de vida-deuda:" in texto:
            datos["seguro_vida"] = texto.split("seguro de vida-deuda:")[1].strip()
            #limpiar signo de dollar y convertirlo en float
            datos["seguro_vida"] = float(re.sub(r'[^\d.]', '', datos["seguro_vida"]))
        elif "gastos notariales:" in texto:
            datos["gastos_notariales"] = texto.split("gastos notariales:")[1].strip()
            datos["gastos_notariales"] = float(re.sub(r'[^\d.]', '', datos["gastos_notariales"]))
        elif "gastos registrales:" in texto:
            datos["gastos_registrales"] = texto.split("gastos registrales:")[1].strip()
            datos["gastos_registrales"] = float(re.sub(r'[^\d.]', '', datos["gastos_registrales"]))
        elif "otros:" in texto:
            datos["otros"] = texto.split("otros:")[1].strip()
            datos["otros"] = float(re.sub(r'[^\d.]', '', datos["otros"]))
        elif "cancelacion credito anterior:" in texto:
            datos["cancelacion_saldo"] = texto.split("cancelacion credito anterior:")[1].strip()
            datos["cancelacion_saldo"] = float(re.sub(r'[^\d.]', '', datos["cancelacion_saldo"]))
        elif "nº credito anterior:" in texto:
            datos["n_credito_cancelar"] = texto.split("nº credito anterior:")[1].strip()
            datos["n_credito_cancelar"] = int(re.sub(r'[^\d.]', '', datos["n_credito_cancelar"]))
        elif "total de descuentos:" in texto:
            datos["total_descuentos"] = texto.split("total de descuentos:")[1].strip()
            datos["total_descuentos"] = float(re.sub(r'[^\d.]', '', datos["total_descuentos"]))
        elif "total entregar al cliente:" in texto:
            datos["desembolso_cliente"] = texto.split("total entregar al cliente:")[1].strip()
            datos["desembolso_cliente"] = float(re.sub(r'[^\d.]', '', datos["desembolso_cliente"]))
        else:
            print(f"Texto no reconocido: {texto}")
            
    print(f"Datos extraídos de Word: {datos}")
    return datos

# Función para extraer datos de archivos Excel
def extraer_datos_excel(ruta_archivo):
    datos = {}
    try:
        workbook = load_workbook(filename=ruta_archivo, data_only=True)
        hoja = workbook.active
        ultima_fila_verde = None
        ultima_celda_amarilla = None
        datos['monto'] = float(re.sub(r'[^\d.]', '', str(hoja.cell(row=2, column=2).value)))
        datos['interes'] = float(re.sub(r'[^\d.]', '', str(hoja.cell(row=3, column=2).value)))
        datos['periodo'] = hoja.cell(row=4, column=2).value.split(" ")[1]
        #si el periodo = Semanal, Quincenal, Mensual, Trimestral, Semestral, Anual usar 1, 3, 4, 5 respectivamente
        if "semanal" in datos['periodo']:
            datos['periodo'] = 1
        elif "quincenal" in datos['periodo']:
            datos['periodo'] = 3
        elif "mensual" in datos['periodo']:
            datos['periodo'] = 4
        elif "trimestral" in datos['periodo']:
            datos['periodo'] = 5
        else:
            datos['periodo'] = 0
        
        datos['plazo'] = int(hoja.cell(row=5, column=2).value.split(" ")[1])
        datos['valor_cuota'] = hoja.cell(row=8, column=3).value
        #si la celda tiene el valor de Seguro devolver tr
        if(hoja.cell(row=7, column=6).value == "Seguro"):
            datos['seguro'] = True
        else:
            datos['seguro'] = False

        verde = 'FFC6E0B4'
        amarillo = 'FFFFFF00'

        #print cell g 10 color
        cell = hoja.cell(row=8, column=1)
        cell_color = cell.fill.start_color.index
        # Recorrer todas las filas en la hoja activa
        celdas_comentadas = []
        for row in hoja.iter_rows():
            for cell in row:
                comentario = cell.comment
                
                if comentario:
                    #remove e txto [Comentario encadenado]\n\nTu versión de Excel te permite leer este comentario encadenado; sin embargo, las ediciones que se apliquen se quitarán si el archivo se abre en una versión más reciente de Excel. Más información: https://go.microsoft.com/fwlink/?linkid=870924\n\nComentario:\n   
                    clean = comentario.text.replace("[Comentario encadenado]\n\nTu versión de Excel te permite leer este comentario encadenado; sin embargo, las ediciones que se apliquen se quitarán si el archivo se abre en una versión más reciente de Excel. Más información: https://go.microsoft.com/fwlink/?linkid=870924\n\nComentario:\n", "")
                    comentario.text = clean
                    
                          
                    
                    if "fecha valor" in comentario.text:
                        fecha = comentario.text.split("valor")[1].strip()
                        comentario.text = fecha
                    elif "Cliente abonó" in comentario.text:
                        fecha = comentario.text.split("el")[1].split("y")[0].strip()
                        comentario.text = fecha
                    
                    # Agregar la celda con comentario y su contenido a la lista
                    celdas_comentadas.append({
                        
                        'celda': cell.coordinate,
                        'comentario': comentario.text,
                        'valor': cell.value,
                        'valor_limpio': float( re.sub(r'[^\d.]', '', str(cell.value)) )
                    })
        #agregar la ultima celda comentada y a datos
        datos["ultima_fecha_pago"] = celdas_comentadas[-1]["comentario"]
        #posiblemente tenga 1 punto y mas texto, dejar solo antes del punto
        datos['ultima_fecha_pago'] = datos["ultima_fecha_pago"].split(".")[0]
        #reformat from 12-12-22 to 2022-12-12
        datos["ultima_fecha_pago"] = datos["ultima_fecha_pago"].replace("/", "-")
        datos["ultima_fecha_pago"] ="20"+ datos["ultima_fecha_pago"][6:] + "-" + datos["ultima_fecha_pago"][3:5] + "-" + datos["ultima_fecha_pago"][:2]
        #quitar espacios y letras si hay con re
        datos["ultima_fecha_pago"] = re.sub(r'[^\d-]', '', datos["ultima_fecha_pago"])


        datos["saldo"] = celdas_comentadas[-1]["valor_limpio"]
    except Exception as e:
        print(f"Error al leer el archivo Excel {ruta_archivo}: {e}")
        
    return datos


# Procesar todas las carpetas en el directorio base
for carpeta_cliente in os.listdir(directorio_base):
    ruta_carpeta = os.path.join(directorio_base, carpeta_cliente)
    if os.path.isdir(ruta_carpeta):
        print(f"Procesando carpeta del cliente: {carpeta_cliente}")
        datos_cliente = {}
        
        # Recorrer los archivos en la carpeta del cliente
        for archivo in os.listdir(ruta_carpeta):
            ruta_archivo = os.path.join(ruta_carpeta, archivo)
            if archivo.startswith("APROBACION_") and archivo.endswith(".docx"):
                datos_cliente.update(extraer_datos_word(ruta_archivo))
            elif archivo.startswith("DESEMBOLSO_") and archivo.endswith(".docx"):
                datos_cliente.update(extraer_datos_word(ruta_archivo))
            elif archivo.startswith("TablaAmortizacion_") and archivo.endswith(".xlsx"):
                datos_cliente.update(extraer_datos_excel(ruta_archivo))
        
        # Guardar los datos agrupados por cliente
        nombre_cliente = carpeta_cliente.strip().lower()
        if nombre_cliente:
            print(f"  Guardando datos del cliente: {nombre_cliente}")
            datos_por_cliente.append(datos_cliente)
        else:
            print("  No se encontró el nombre del cliente en los datos extraídos")

# Guardar los datos en un archivo JSON
ruta_json = "datos_clientes.json"
print(f"Guardando datos en {ruta_json}")
with open(ruta_json, 'w', encoding='utf-8') as archivo_json:
    json.dump(datos_por_cliente, archivo_json, indent=4, ensure_ascii=False)

print(f"Datos guardados en {ruta_json}")

for cliente, datos in datos_por_cliente:
    print(f"Cliente: {cliente}")
    for clave, valor in datos.items():
        print(f"  {clave}: {valor}")
    print()
